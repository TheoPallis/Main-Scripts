import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, PatternFill

def format_df(file) :
    workbook = openpyxl.load_workbook(file)
    for worksheet in workbook:
        font = Font(color='FFFFFF', bold=True)
        fill = PatternFill(start_color='5552A2', end_color='5552A2', fill_type='solid')
        for cell in worksheet[1]:
            cell.font = font
            cell.fill = fill
        for column in worksheet.columns:
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = 30      
    workbook.save(file)
